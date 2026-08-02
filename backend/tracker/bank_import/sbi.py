"""SBI Excel statement decrypt + row extraction for bulk import preview."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Optional

import msoffcrypto
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# OLE compound-file magic (password-protected Office wrappers).
_OLE_CFB_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

_HEADER_SCAN_MAX_ROWS = 60

SBI_HEADER_ALIASES = {
    "date": "date",
    "details": "details",
    "ref no/cheque no": "ref",
    "ref no./cheque no": "ref",
    "ref no /cheque no": "ref",
    "debit": "debit",
    "credit": "credit",
    "balance": "balance",
}


class BankStatementPasswordError(ValueError):
    """Raised when a password is missing or incorrect for an encrypted workbook."""


class BankStatementParseError(ValueError):
    """Raised when the workbook cannot be interpreted as an SBI statement."""


@dataclass
class SbiRawRow:
    source_row: int
    transaction_date: str
    description: str
    amount: str
    is_debit: str


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"\s+", " ", text)


def decrypt_office_file(content: bytes, password: Optional[str]) -> bytes:
    """
    Return workbook bytes, decrypting when Office encryption is present.

    Unencrypted files (plain .xlsx zip, etc.) are returned unchanged.
    """
    if not content.startswith(_OLE_CFB_MAGIC):
        return content

    try:
        office = msoffcrypto.OfficeFile(io.BytesIO(content))
    except Exception as exc:
        raise BankStatementParseError(
            "Could not open Excel file. Ensure it is a valid .xlsx statement."
        ) from exc

    if not office.is_encrypted():
        return content

    if not (password or "").strip():
        raise BankStatementPasswordError("Password required to open this file.")

    try:
        decrypted = io.BytesIO()
        office.load_key(password=password.strip())
        office.decrypt(decrypted)
        return decrypted.getvalue()
    except Exception as exc:
        raise BankStatementPasswordError("Incorrect password.") from exc


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text.replace("\r", " ").replace("\n", " "))
    return text.strip()


def _to_iso_date(value: object) -> str:
    """Normalize SBI date cells to YYYY-MM-DD for DatePicker / preview validation."""
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # Excel serial date (common when the cell is numeric + date-formatted).
        from openpyxl.utils.datetime import from_excel

        converted = from_excel(value)
        if isinstance(converted, datetime):
            return converted.date().isoformat()
        if isinstance(converted, date):
            return converted.isoformat()
        return ""

    text = _cell_text(value)
    if not text:
        return ""

    # Already ISO.
    if len(text) >= 10 and text[4] == "-" and text[7] == "-":
        try:
            return date.fromisoformat(text[:10]).isoformat()
        except ValueError:
            pass

    for sep in ("/", "-", "."):
        if sep not in text:
            continue
        parts = text.split(sep)
        if len(parts) != 3:
            continue
        try:
            p1, p2, p3 = (int(p) for p in parts)
        except ValueError:
            continue
        # YYYY-MM-DD / YYYY/MM/DD
        if p1 >= 1000:
            year, month, day = p1, p2, p3
        else:
            # SBI statements use DD/MM/YYYY (India).
            day, month, year = p1, p2, p3
            if year < 100:
                year = 2000 + year if year <= 68 else 1900 + year
        try:
            return date(year, month, day).isoformat()
        except ValueError:
            continue
    return text


def _find_header_map(ws: Worksheet) -> tuple[int, dict[str, int]]:
    for row_idx, row in enumerate(
        ws.iter_rows(max_row=_HEADER_SCAN_MAX_ROWS, values_only=True),
        start=1,
    ):
        mapping: dict[str, int] = {}
        for col_idx, cell in enumerate(row or (), start=1):
            key = SBI_HEADER_ALIASES.get(_normalize_header(cell))
            if key:
                mapping[key] = col_idx
        if {"date", "details", "debit", "credit"}.issubset(mapping):
            return row_idx, mapping
    raise BankStatementParseError(
        "Could not find SBI statement headers (Date, Details, Debit, Credit)."
    )


def _is_summary_or_blank(date_text: str, details: str, debit: str, credit: str) -> bool:
    if not date_text and not details and not debit and not credit:
        return True
    blob = f"{date_text} {details}".lower()
    if "statement summary" in blob:
        return True
    if details.lower().startswith("brought forward"):
        return True
    if "dr count" in blob or "total debits" in blob:
        return True
    return False


def extract_sbi_rows(content: bytes, password: Optional[str] = None) -> list[SbiRawRow]:
    """Decrypt (if needed) and extract transaction rows from an SBI .xlsx statement."""
    workbook_bytes = decrypt_office_file(content, password)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise BankStatementParseError(
            "Could not read Excel file. Ensure it is a valid .xlsx statement."
        ) from exc

    try:
        ws = wb[wb.sheetnames[0]]
        header_row, colmap = _find_header_map(ws)
        rows: list[SbiRawRow] = []

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if row is None:
                continue

            def col(name: str) -> str:
                idx = colmap.get(name)
                if not idx or idx > len(row):
                    return ""
                return _cell_text(row[idx - 1])

            date_idx = colmap.get("date")
            if not date_idx or date_idx > len(row):
                date_text = ""
            else:
                date_text = _to_iso_date(row[date_idx - 1])
            details = col("details")
            debit = col("debit").replace(",", "")
            credit = col("credit").replace(",", "")

            if _is_summary_or_blank(date_text, details, debit, credit):
                # Summary block follows the txn table; stop once we have rows.
                if rows and (not date_text or "summary" in details.lower()):
                    break
                continue

            if not date_text:
                continue

            has_debit = bool(debit)
            has_credit = bool(credit)
            if has_debit == has_credit:
                continue

            rows.append(
                SbiRawRow(
                    source_row=row_idx,
                    transaction_date=date_text,
                    description=details,
                    amount=debit if has_debit else credit,
                    is_debit="true" if has_debit else "false",
                )
            )
    finally:
        wb.close()

    if not rows:
        raise BankStatementParseError("No transactions found in the SBI statement.")

    return rows
